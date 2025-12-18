"""
Application web Flask pour le priceur d'options Long Straddle
Interface web moderne pour analyser les stratégies d'options
"""

from flask import Flask, render_template, request, jsonify, send_file
import json
import os
import numpy as np
from io import BytesIO
from datetime import datetime
from src.strategies.long_straddle import LongStraddle
from src.strategies.long_strangle import LongStrangle
from src.strategies.iron_condor import IronCondor
from src.models.black_scholes import Call, Put
from src.utils.market_data import get_ticker_info, validate_ticker, get_historical_volatility
from src.utils.monte_carlo import MonteCarloAnalysis
from src.utils.backtesting import Backtester

# Configuration
OUTPUT_DIR = 'output'
os.makedirs(OUTPUT_DIR, exist_ok=True)

app = Flask(__name__, 
            template_folder='web/templates',
            static_folder='web/static')

# Stocker l'historique des analyses (en mémoire pour la démo)
analysis_history = []


@app.route('/')
def index():
    """Page d'accueil"""
    return render_template('index.html')


@app.route('/api/validate_ticker', methods=['POST'])
def validate_ticker_api():
    """API pour valider un ticker"""
    data = request.get_json()
    ticker = data.get('ticker', '').strip().upper()
    
    if not ticker:
        return jsonify({'valid': False, 'error': 'Ticker vide'})
    
    is_valid = validate_ticker(ticker)
    
    if is_valid:
        try:
            ticker_info = get_ticker_info(ticker)
            return jsonify({
                'valid': True,
                'ticker_info': ticker_info
            })
        except Exception as e:
            return jsonify({'valid': False, 'error': str(e)})
    else:
        return jsonify({'valid': False, 'error': 'Ticker invalide ou non trouvé'})


@app.route('/api/calculate_straddle', methods=['POST'])
def calculate_straddle():
    """API pour calculer le straddle"""
    data = request.get_json()
    
    ticker = data.get('ticker', '').strip().upper()
    days = int(data.get('days', 30))
    custom_strike = data.get('strike')
    
    if custom_strike:
        try:
            custom_strike = float(custom_strike)
        except:
            custom_strike = None
    
    try:
        # Créer le straddle
        straddle = LongStraddle.from_ticker(ticker, K=custom_strike, days_to_expiry=days)
        
        # Générer le résumé
        summary = straddle.summary()
        
        # Calculer les scénarios de profit
        scenarios = []
        price_variations = list(range(-50, 51, 5))  # -50% à +50% par pas de 5%
        
        for pct in price_variations:
            final_price = straddle.S * (1 + pct/100)
            profit = straddle.profit_at_expiry(final_price)
            payoff = straddle.payoff_at_expiry(final_price)
            
            scenarios.append({
                'price_change_pct': pct,
                'final_price': final_price,
                'payoff': payoff,
                'profit': profit
            })
        
        return jsonify({
            'success': True,
            'summary': summary,
            'scenarios': scenarios
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        })


@app.route('/api/ticker_info/<ticker>')
def ticker_info_api(ticker):
    """API pour obtenir les informations d'un ticker"""
    try:
        info = get_ticker_info(ticker.upper())
        return jsonify({'success': True, 'info': info})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/compare_strategies', methods=['POST'])
def compare_strategies():
    """API pour comparer plusieurs configurations de straddle"""
    data = request.get_json()
    ticker = data.get('ticker', '').strip().upper()
    
    try:
        # Récupérer le prix spot
        from src.utils.market_data import get_market_data
        S, sigma, r = get_market_data(ticker)
        
        results = []
        
        # Générer différentes configurations
        days_options = [7, 14, 30, 60, 90]
        strike_offsets = [-0.05, 0, 0.05]  # -5%, ATM, +5%
        
        for days in days_options:
            for offset in strike_offsets:
                strike = S * (1 + offset)
                straddle = LongStraddle.from_ticker(ticker, K=strike, days_to_expiry=days)
                summary = straddle.summary()
                
                results.append({
                    'days': days,
                    'strike': strike,
                    'strike_type': 'ATM' if offset == 0 else f'{int(offset*100):+d}%',
                    'cost': summary['total_cost'],
                    'break_even_move': summary['break_even_move_pct'],
                    'theta': summary['greeks']['theta'],
                    'vega': summary['greeks']['vega']
                })
        
        return jsonify({'success': True, 'comparisons': results})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/greeks_sensitivity', methods=['POST'])
def greeks_sensitivity():
    """API pour calculer la sensibilité des Greeks"""
    data = request.get_json()
    
    ticker = data.get('ticker', '').strip().upper()
    days = int(data.get('days', 30))
    custom_strike = data.get('strike')
    
    if custom_strike:
        try:
            custom_strike = float(custom_strike)
        except:
            custom_strike = None
    
    try:
        from src.utils.market_data import get_market_data
        S, sigma, r = get_market_data(ticker)
        
        if custom_strike is None:
            custom_strike = S
        
        T = days / 365.0
        
        # Sensibilité à la volatilité
        vol_range = np.linspace(sigma * 0.5, sigma * 1.5, 20)
        vol_sensitivity = []
        
        for vol in vol_range:
            straddle = LongStraddle(S, custom_strike, T, r, vol)
            price = straddle.price()
            greeks = straddle.greeks()
            
            vol_sensitivity.append({
                'volatility': vol * 100,
                'price': price,
                'vega': greeks['vega']
            })
        
        # Sensibilité au temps
        time_range = np.linspace(1, days, min(days, 20))
        time_sensitivity = []
        
        for d in time_range:
            T_temp = d / 365.0
            straddle = LongStraddle(S, custom_strike, T_temp, r, sigma)
            price = straddle.price()
            greeks = straddle.greeks()
            
            time_sensitivity.append({
                'days': int(d),
                'price': price,
                'theta': greeks['theta']
            })
        
        # Sensibilité au prix spot
        spot_range = np.linspace(S * 0.8, S * 1.2, 30)
        spot_sensitivity = []
        
        for spot in spot_range:
            straddle = LongStraddle(spot, custom_strike, T, r, sigma)
            price = straddle.price()
            greeks = straddle.greeks()
            
            spot_sensitivity.append({
                'spot_price': spot,
                'price': price,
                'delta': greeks['delta'],
                'gamma': greeks['gamma']
            })
        
        return jsonify({
            'success': True,
            'volatility_sensitivity': vol_sensitivity,
            'time_sensitivity': time_sensitivity,
            'spot_sensitivity': spot_sensitivity
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/heatmap_data', methods=['POST'])
def heatmap_data():
    """API pour générer une heatmap de profit"""
    data = request.get_json()
    
    ticker = data.get('ticker', '').strip().upper()
    custom_strike = data.get('strike')
    
    if custom_strike:
        try:
            custom_strike = float(custom_strike)
        except:
            custom_strike = None
    
    try:
        from src.utils.market_data import get_market_data
        S, sigma, r = get_market_data(ticker)
        
        if custom_strike is None:
            custom_strike = S
        
        # Générer la heatmap: jours vs variation de prix
        days_range = [7, 14, 21, 30, 45, 60, 90]
        price_changes = list(range(-30, 31, 5))  # -30% à +30%
        
        heatmap = []
        
        for days in days_range:
            row = []
            T = days / 365.0
            
            for pct in price_changes:
                final_price = S * (1 + pct/100)
                straddle = LongStraddle(S, custom_strike, T, r, sigma)
                profit = straddle.profit_at_expiry(final_price)
                
                row.append({
                    'days': days,
                    'price_change': pct,
                    'profit': profit
                })
            
            heatmap.append(row)
        
        return jsonify({
            'success': True,
            'heatmap': heatmap,
            'days_range': days_range,
            'price_changes': price_changes
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/export_json', methods=['POST'])
def export_json():
    """API pour exporter l'analyse en JSON"""
    data = request.get_json()
    
    # Ajouter timestamp
    data['timestamp'] = datetime.now().isoformat()
    
    # Ajouter à l'historique
    analysis_history.append(data)
    
    return jsonify({
        'success': True,
        'data': data
    })


@app.route('/api/history')
def get_history():
    """API pour récupérer l'historique des analyses"""
    return jsonify({
        'success': True,
        'history': analysis_history[-10:]  # 10 dernières analyses
    })


@app.route('/api/implied_volatility', methods=['POST'])
def implied_volatility():
    """API pour calculer la volatilité implicite historique"""
    data = request.get_json()
    ticker = data.get('ticker', '').strip().upper()
    
    try:
        # Calculer la volatilité sur différentes périodes
        periods = [30, 60, 90, 180, 252]
        vol_data = []
        
        for period in periods:
            vol = get_historical_volatility(ticker, period)
            vol_data.append({
                'period': period,
                'period_label': f'{period}d',
                'volatility': vol * 100
            })
        
        return jsonify({
            'success': True,
            'volatility_data': vol_data
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/export_pdf', methods=['POST'])
def export_pdf():
    """API pour exporter l'analyse en PDF"""
    data = request.get_json()
    
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import inch
        
        filename = f"{OUTPUT_DIR}/straddle_{data.get('ticker', 'UNKNOWN')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        doc = SimpleDocTemplate(filename, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=24, textColor=colors.HexColor('#3b82f6'))
        story.append(Paragraph(f"Long Straddle Analysis", title_style))
        story.append(Spacer(1, 0.3*inch))
        
        summary = data.get('summary', {})
        ticker_data = [
            ['Ticker', data.get('ticker', 'N/A')],
            ['Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Prix Spot', f"${summary.get('spot_price', 0):.2f}"],
            ['Strike', f"${summary.get('strike', 0):.2f}"],
            ['Échéance', f"{summary.get('time_to_expiry_days', 0)} jours"],
            ['Coût Total', f"${summary.get('total_cost', 0):.2f}"],
        ]
        
        t1 = Table(ticker_data, colWidths=[2*inch, 3*inch])
        t1.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e0e7ff')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(t1)
        doc.build(story)
        
        return send_file(filename, as_attachment=True, download_name=os.path.basename(filename))
    except ImportError:
        return jsonify({'success': False, 'error': 'reportlab non installé. Installez avec: pip install reportlab'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/export_excel', methods=['POST'])
def export_excel():
    """API pour exporter l'analyse en Excel"""
    data = request.get_json()
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Straddle Analysis"
        
        ws['A1'] = "Options Pricer - Long Straddle"
        ws['A1'].font = Font(bold=True, size=16, color="3b82f6")
        
        summary = data.get('summary', {})
        row = 3
        for label, key in [
            ['Ticker', 'ticker'],
            ['Prix Spot', 'spot_price'],
            ['Strike', 'strike'],
            ['Coût Total', 'total_cost'],
        ]:
            ws.cell(row, 1, label).font = Font(bold=True)
            if key == 'ticker':
                ws.cell(row, 2, data.get(key, 'N/A'))
            else:
                val = summary.get(key, 0)
                ws.cell(row, 2, f"${val:.2f}" if isinstance(val, (int, float)) else val)
            row += 1
        
        filename = f"{OUTPUT_DIR}/straddle_{data.get('ticker', 'UNKNOWN')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        
        return send_file(filename, as_attachment=True, download_name=os.path.basename(filename))
    except ImportError:
        return jsonify({'success': False, 'error': 'openpyxl non installé. Installez avec: pip install openpyxl'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/export_csv', methods=['POST'])
def export_csv():
    """API pour exporter les scénarios en CSV"""
    data = request.get_json()
    
    try:
        import csv
        
        filename = f"{OUTPUT_DIR}/scenarios_{data.get('ticker', 'UNKNOWN')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        with open(filename, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(['Variation %', 'Prix Final', 'Payoff', 'Profit/Perte'])
            
            for scenario in data.get('scenarios', []):
                writer.writerow([
                    f"{scenario['price_change_pct']}%",
                    f"${scenario['final_price']:.2f}",
                    f"${scenario['payoff']:.2f}",
                    f"${scenario['profit']:.2f}"
                ])
        
        return send_file(filename, as_attachment=True, download_name=os.path.basename(filename))
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/compare_multi_strategies', methods=['POST'])
def compare_multi_strategies():
    """Compare Long Straddle, Long Strangle et Iron Condor"""
    data = request.get_json()
    
    try:
        ticker = data.get('ticker')
        days = int(data.get('days', 30))
        
        info = get_ticker_info(ticker)
        spot_price = info['current_price']
        
        # Créer les 3 stratégies
        straddle = LongStraddle.from_ticker(ticker, days)
        strangle = LongStrangle.from_ticker(ticker, days, otm_percent=0.05)
        condor = IronCondor.from_ticker(ticker, days)
        
        # Générer des scénarios de prix
        price_range = np.linspace(spot_price * 0.7, spot_price * 1.3, 100)
        
        strategies_data = {
            'straddle': {
                'summary': straddle.summary(),
                'payoffs': [straddle.profit_at_expiry(p) for p in price_range]
            },
            'strangle': {
                'summary': strangle.summary(),
                'payoffs': [strangle.profit_at_expiry(p) for p in price_range]
            },
            'iron_condor': {
                'summary': condor.summary(),
                'payoffs': [condor.profit_at_expiry(p) for p in price_range]
            },
            'price_range': price_range.tolist(),
            'spot_price': spot_price
        }
        
        return jsonify({
            'success': True,
            'strategies': strategies_data
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/monte_carlo', methods=['POST'])
def monte_carlo_analysis():
    """Analyse Monte Carlo pour probabilité de profit"""
    data = request.get_json()
    
    try:
        ticker = data.get('ticker')
        days = int(data.get('days', 30))
        strategy_type = data.get('strategy', 'straddle')
        num_simulations = int(data.get('simulations', 10000))
        
        info = get_ticker_info(ticker)
        spot_price = info['current_price']
        volatility = info.get('implied_volatility', 0.3)
        
        # Créer la stratégie
        if strategy_type == 'straddle':
            strategy = LongStraddle.from_ticker(ticker, days)
        elif strategy_type == 'strangle':
            strategy = LongStrangle.from_ticker(ticker, days)
        elif strategy_type == 'iron_condor':
            strategy = IronCondor.from_ticker(ticker, days)
        else:
            return jsonify({'success': False, 'error': 'Stratégie inconnue'})
        
        # Monte Carlo
        mc = MonteCarloAnalysis(spot_price, volatility)
        mc_result = mc.probability_of_profit(
            strategy.profit_at_expiry,
            days / 365.0,
            num_simulations
        )
        
        # VaR analysis
        var_result = mc.value_at_risk(
            strategy.profit_at_expiry,
            days / 365.0,
            confidence_level=0.95,
            num_simulations=num_simulations
        )
        
        # Break-even probability
        lower_be, upper_be = strategy.break_even_points()
        be_analysis = mc.breakeven_probability_analysis(
            (lower_be, upper_be),
            days / 365.0,
            num_simulations
        )
        
        return jsonify({
            'success': True,
            'monte_carlo': mc_result,
            'value_at_risk': var_result,
            'breakeven_analysis': be_analysis
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/backtest', methods=['POST'])
def backtest_strategy():
    """Backtest historique d'une stratégie"""
    data = request.get_json()
    
    try:
        ticker = data.get('ticker')
        strategy_type = data.get('strategy', 'straddle')
        holding_days = int(data.get('holding_days', 30))
        
        # Période de backtest (dernière année)
        end_date = datetime.now().strftime('%Y-%m-%d')
        start_date = (datetime.now().replace(year=datetime.now().year - 1)).strftime('%Y-%m-%d')
        
        backtester = Backtester(ticker, start_date, end_date)
        
        # Déterminer la stratégie
        if strategy_type == 'straddle':
            from src.strategies.long_straddle import LongStraddle
            result = backtester.backtest_strategy(LongStraddle, holding_period_days=holding_days)
        elif strategy_type == 'strangle':
            from src.strategies.long_strangle import LongStrangle
            result = backtester.backtest_strategy(LongStrangle, holding_period_days=holding_days, otm_percent=0.05)
        else:
            return jsonify({'success': False, 'error': 'Stratégie non supportée pour le backtest'})
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/glossary', methods=['GET'])
def get_glossary():
    """Retourne le glossaire complet"""
    try:
        with open('web/static/glossary.json', 'r', encoding='utf-8') as f:
            glossary = json.load(f)
        return jsonify({'success': True, 'glossary': glossary})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


if __name__ == '__main__':
    print("\n" + "="*70)
    print("  OPTIONS PRICER - Interface Web Améliorée")
    print("="*70)
    print("\n  🌐 Serveur démarré sur: http://127.0.0.1:5003")
    print("  📊 Accédez à l'interface web dans votre navigateur")
    print(f"  📁 Exports sauvegardés dans: {os.path.abspath(OUTPUT_DIR)}/")
    print("  📚 Nouvelles fonctionnalités: Monte Carlo, Backtesting, Multi-stratégies")
    print("\n  Appuyez sur Ctrl+C pour arrêter le serveur\n")
    
    app.run(debug=True, host='0.0.0.0', port=5003)
